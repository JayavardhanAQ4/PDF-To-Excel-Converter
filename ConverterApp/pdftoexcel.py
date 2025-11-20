import os
from io import BytesIO
import re
from datetime import datetime
from collections import defaultdict
import pdfplumber
import fitz  # PyMuPDF
import pandas as pd
from flask import Flask, render_template, request, redirect, send_file, flash
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

try:
    import pytesseract
    from PIL import Image
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False

# ------------------------
# Config / Flask setup
# ------------------------
app = Flask(__name__, template_folder="templates")
app.secret_key = "change_this_secret_in_prod"

UPLOAD_FOLDER = "uploads"
ALLOWED_EXTENSIONS = {"pdf"}
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


def allowed(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


# ---------------------------
# Helpers: normalize / style / autofit / numeric parsing
# ---------------------------
def normalize_cell_text(v):
    """Make table cell text Excel-friendly: strip, replace many spaces/newlines with single space"""
    if v is None:
        return ""
    s = str(v)
    # replace newlines inside cells with space, collapse multiple spaces/tabs
    s = " ".join(s.replace("\r", " ").replace("\n", " ").split())
    return s


def _apply_table_style(ws, start_row, start_col, nrows, ncols, header_rows=1):
    """Apply simple borders and header bold on the inserted table range"""
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # header bold for header_rows
    for hr in range(header_rows):
        row_idx = start_row + hr
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=row_idx, column=c)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrapText=True, vertical="top")

    # borders for whole table
    for r in range(start_row, start_row + nrows):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = Alignment(wrapText=True, vertical="top")


def _auto_fit_columns(ws, min_col=1, max_col=30, max_width=60):
    # naive auto-fit using current content
    for c in range(min_col, max_col + 1):
        col_letter = ws.cell(row=1, column=c).column_letter
        max_len = 0
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        if max_len == 0:
            continue
        ws.column_dimensions[col_letter].width = min(max_width, max(10, int(max_len * 1.1)))


_NUMERIC_RE = re.compile(r"^[\+\-]?\d{1,3}([,.\s]\d{3})*([.,]\d+)?%?$")
_PERCENT_RE = re.compile(r"^([\+\-]?\d+[.,]?\d*)\s*%$")
_DATE_PATTERNS = [
    r"\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b",
    r"\b\d{4}[/-]\d{1,2}[/-]\d{1,2}\b",
    r"\b(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{1,2},?\s+\d{4}\b",
    r"\b\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{4}\b"
]


def try_parse_date(s):
    """Try to parse date strings into datetime objects"""
    if not s or not isinstance(s, str):
        return (None, False)
    
    s = s.strip()
    date_formats = [
        "%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%d", "%Y/%m/%d",
        "%m-%d-%Y", "%d-%m-%Y", "%m/%d/%y", "%d/%m/%y",
        "%B %d, %Y", "%b %d, %Y", "%d %B %Y", "%d %b %Y"
    ]
    
    for fmt in date_formats:
        try:
            dt = datetime.strptime(s, fmt)
            return (dt, True)
        except:
            continue
    return (None, False)


def try_parse_number(s):
    """
    Try to parse a cell string into int/float or percent.
    Returns (value, is_number_bool).
    Handles: "1,234.56", "1.234,56", "1234,56", "1 234", "12%", "(1234)" -> negative,
    currency symbols removed, spaces trimmed.
    """
    if s is None:
        return (None, False)
    raw = str(s).strip()
    if raw == "":
        return (None, False)

    # remove common currency symbols and non-breaking spaces
    cleaned = raw.replace("\u00A0", " ").strip()
    cleaned = re.sub(r"[£€$₹¥]", "", cleaned)

    # handle parentheses as negative e.g. (1,234.56)
    negative = False
    if cleaned.startswith("(") and cleaned.endswith(")"):
        negative = True
        cleaned = cleaned[1:-1].strip()

    # detect percent
    m_pct = _PERCENT_RE.match(cleaned)
    if m_pct:
        numpart = m_pct.group(1)
        numpart = numpart.replace(" ", "")
        numpart = numpart.replace(",", ".") if numpart.count(",") == 1 and numpart.count(".") == 0 else numpart.replace(",", "")
        try:
            val = float(numpart)
            if negative:
                val = -val
            return (val / 100.0, True)
        except Exception:
            return (None, False)

    # remove thousands separators intelligently:
    # if both '.' and ',' occur, assume last one is decimal separator.
    if cleaned.count(",") and cleaned.count("."):
        # assume last symbol is decimal separator
        if cleaned.rfind(",") > cleaned.rfind("."):
            # comma decimal -> remove dots
            cleaned = cleaned.replace(".", "")
            cleaned = cleaned.replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    else:
        # if only commas and they separate groups (like 1,234) -> remove them
        if cleaned.count(",") and len(cleaned.split(",")) > 1:
            # if last chunk length is 3 it's likely thousands sep
            if len(cleaned.split(",")[-1]) == 3:
                cleaned = cleaned.replace(",", "")
            else:
                cleaned = cleaned.replace(",", ".")
        # replace spaces as thousands sep
        cleaned = cleaned.replace(" ", "")

    # final attempt
    try:
        if cleaned == "":
            return (None, False)
        if "." in cleaned:
            val = float(cleaned)
        else:
            val = int(cleaned)
        if negative:
            val = -val
        return (val, True)
    except Exception:
        return (None, False)


def detect_merged_cells(table):
    """Detect cells that should be merged based on repeated values - DISABLED to prevent scroll issues"""
    # Merged cells can cause Excel navigation issues, so we disable this feature
    return []


def repair_table_structure(table):
    """Fix common table issues like ragged rows"""
    if not table:
        return table
    
    max_cols = max(len(row) for row in table)
    repaired = []
    for row in table:
        new_row = list(row) + [""] * (max_cols - len(row))
        repaired.append(new_row)
    
    return repaired


def detect_column_positions(x_positions, page_width, tolerance=30):
    """Detect column boundaries from x-positions"""
    if not x_positions:
        return [0, page_width]
    
    # Cluster x-positions to find column starts
    x_positions = sorted(set(x_positions))
    columns = [0]
    
    for i, x in enumerate(x_positions):
        if i == 0 or x - x_positions[i-1] > tolerance:
            if x > tolerance:  # Not too close to left edge
                columns.append(x)
    
    columns.append(page_width)
    return sorted(set(columns))


def assign_to_column(x_pos, column_boundaries):
    """Assign x-position to nearest column"""
    for i in range(len(column_boundaries) - 1):
        if column_boundaries[i] <= x_pos < column_boundaries[i + 1]:
            return i
    return len(column_boundaries) - 2


def detect_text_alignment(words, page_width):
    """Detect if text is left, center, or right aligned"""
    if not words:
        return 'left'
    
    first_x = words[0]['x0']
    last_x = words[-1]['x1'] if 'x1' in words[-1] else words[-1]['x0']
    
    # Calculate position relative to page
    left_margin = first_x
    right_margin = page_width - last_x
    
    if abs(left_margin - right_margin) < 50:  # Centered
        return 'center'
    elif left_margin > page_width * 0.6:  # Right aligned
        return 'right'
    else:
        return 'left'


def calculate_optimal_width(values, min_width=10, max_width=50):
    """Calculate optimal column width based on content"""
    if not values:
        return min_width
    
    max_len = 0
    for val in values:
        if val is not None:
            str_val = str(val)
            # Better width calculation
            if isinstance(val, (int, float)):
                max_len = max(max_len, len(str_val) * 1.1)
            elif isinstance(val, datetime):
                max_len = max(max_len, 12)  # Standard date width
            else:
                # Count actual character width more accurately
                max_len = max(max_len, len(str_val) * 1.15)
    
    return min(max_width, max(min_width, int(max_len) + 3))


def detect_multi_column_layout(words, page_width):
    """Detect if page has multi-column layout (like newspapers)"""
    if not words or len(words) < 10:
        return False, []
    
    # Group words by x-position
    x_positions = [w.get('x0', 0) for w in words]
    mid_point = page_width / 2
    
    left_col = sum(1 for x in x_positions if x < mid_point * 0.6)
    right_col = sum(1 for x in x_positions if x > mid_point * 1.4)
    
    # If significant words on both sides, it's multi-column
    if left_col > 20 and right_col > 20:
        return True, [mid_point]
    return False, []


def extract_form_fields(pdf_path):
    """Extract form fields from PDF (like Adobe does)"""
    form_data = []
    try:
        doc = fitz.open(pdf_path)
        for page_num, page in enumerate(doc, 1):
            if page.first_widget:
                widget = page.first_widget
                while widget:
                    field_info = {
                        'page': page_num,
                        'name': widget.field_name or 'Unnamed',
                        'value': widget.field_value or '',
                        'type': widget.field_type_string or 'text'
                    }
                    form_data.append(field_info)
                    widget = widget.next
        doc.close()
    except Exception:
        pass
    return form_data


def extract_images_from_page(page, page_num):
    """Extract images and their positions from PDF page"""
    images = []
    try:
        for img_index, img in enumerate(page.images):
            try:
                x0, y0 = img.get('x0', 0), img.get('top', 0)
                images.append({
                    'page': page_num,
                    'x': x0,
                    'y': y0,
                    'width': img.get('width', 0),
                    'height': img.get('height', 0)
                })
            except:
                continue
    except:
        pass
    return images


def perform_ocr_on_page(pdf_path, page_num):
    """Perform OCR on scanned PDF pages (Adobe-like feature)"""
    if not OCR_AVAILABLE:
        return None
    
    try:
        doc = fitz.open(pdf_path)
        page = doc[page_num - 1]
        pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))  # 2x zoom for better OCR
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        text = pytesseract.image_to_string(img)
        doc.close()
        return text
    except Exception:
        return None


def detect_table_structure_advanced(page):
    """Advanced table detection using multiple strategies"""
    tables = []
    
    # Strategy 1: Line-based detection (best for bordered tables)
    try:
        line_tables = page.find_tables({
            "vertical_strategy": "lines",
            "horizontal_strategy": "lines",
            "intersection_tolerance": 5,
            "snap_tolerance": 5,
            "join_tolerance": 5,
            "edge_min_length": 10,
        })
        for t in line_tables:
            tables.append(('lines', t))
    except:
        pass
    
    # Strategy 2: Text-based detection (for tables without borders)
    try:
        text_tables = page.find_tables({
            "vertical_strategy": "text",
            "horizontal_strategy": "text",
            "intersection_tolerance": 10,
            "text_tolerance": 5,
            "text_x_tolerance": 5,
            "text_y_tolerance": 5,
        })
        for t in text_tables:
            # Avoid duplicates
            is_duplicate = False
            for existing_type, existing_table in tables:
                if abs(t.bbox[0] - existing_table.bbox[0]) < 10:
                    is_duplicate = True
                    break
            if not is_duplicate:
                tables.append(('text', t))
    except:
        pass
    
    # Strategy 3: Explicit lines detection
    try:
        explicit_tables = page.find_tables({
            "vertical_strategy": "explicit",
            "horizontal_strategy": "explicit",
            "explicit_vertical_lines": page.curves + page.edges,
            "explicit_horizontal_lines": page.curves + page.edges,
        })
        for t in explicit_tables:
            is_duplicate = False
            for existing_type, existing_table in tables:
                if abs(t.bbox[0] - existing_table.bbox[0]) < 10:
                    is_duplicate = True
                    break
            if not is_duplicate:
                tables.append(('explicit', t))
    except:
        pass
    
    return tables


# ---------------------------
# Extract LOGO (safer filename)
# ---------------------------
def extract_logo(pdf_path):
    try:
        doc = fitz.open(pdf_path)
    except Exception:
        return None

    imgs = []
    for page in doc[:2]:
        for img in page.get_images(full=True):
            xref = img[0]
            info = doc.extract_image(xref)
            imgs.append(info)
    doc.close()

    if not imgs:
        return None

    best = max(imgs, key=lambda x: x.get("width", 0) * x.get("height", 0))
    out = os.path.splitext(os.path.basename(pdf_path))[0] + "_logo.png"
    try:
        with open(out, "wb") as f:
            f.write(best["image"])
        return out
    except Exception:
        return None


# ---------------------------
# Improved extraction: text blocks + tables with y-coordinates + page height
# ---------------------------
def extract_pdf_content(pdf_path):
    """
    Returns pages with spatial positioning preserved
    """
    pages_out = []
    with pdfplumber.open(pdf_path) as pdf:
        for pno, page in enumerate(pdf.pages, start=1):
            blocks = []

            page_height = getattr(page, "height", None)
            page_width = getattr(page, "width", 612.0)
            if page_height is None:
                try:
                    doc = fitz.open(pdf_path)
                    page_height = doc[pno-1].rect.height
                    page_width = doc[pno-1].rect.width
                    doc.close()
                except Exception:
                    page_height = 792.0
                    page_width = 612.0

            # 1) Extract text with spatial positioning
            try:
                words = page.extract_words(use_text_flow=False, x_tolerance=2, y_tolerance=2)
            except Exception:
                try:
                    words = page.extract_words()
                except:
                    words = []

            # Group words by line AND horizontal position (column detection)
            lines = {}
            for w in words:
                top = round(float(w.get("top", 0)) * 2) / 2
                x0 = float(w.get("x0", 0))
                x1 = float(w.get("x1", 0))
                lines.setdefault(top, []).append({
                    'x0': x0,
                    'x1': x1,
                    'text': w["text"],
                    'width': x1 - x0
                })
            
            # Detect column positions across the page
            all_x_positions = []
            for words_in_line in lines.values():
                for w in words_in_line:
                    all_x_positions.append(w['x0'])
            
            # Create column boundaries (divide page into logical columns)
            column_boundaries = detect_column_positions(all_x_positions, page_width)
            
            for top, words_line in sorted(lines.items(), key=lambda x: x[0]):
                words_line.sort(key=lambda x: x['x0'])
                
                # Map words to columns for proper alignment
                if len(column_boundaries) > 1:
                    # Multi-column layout
                    text_with_position = []
                    for w in words_line:
                        col_idx = assign_to_column(w['x0'], column_boundaries)
                        text_with_position.append((col_idx, w['x0'], w['text']))
                    
                    blocks.append({
                        "type": "text",
                        "y": float(top),
                        "content": words_line,
                        "columns": text_with_position,
                        "alignment": detect_text_alignment(words_line, page_width)
                    })
                else:
                    # Single column - preserve spacing
                    text_line = " ".join(w['text'] for w in words_line)
                    if text_line.strip():
                        blocks.append({
                            "type": "text",
                            "y": float(top),
                            "content": text_line,
                            "x_start": words_line[0]['x0'] if words_line else 0,
                            "alignment": detect_text_alignment(words_line, page_width)
                        })

            # 2) Advanced table extraction with multiple strategies
            detected_tables = detect_table_structure_advanced(page)
            
            for strategy, t in detected_tables:
                try:
                    table_data = t.extract()
                    bbox = t.bbox
                    y_mid = (bbox[1] + bbox[3]) / 2.0
                    
                    # Normalize and repair table structure
                    table_norm = [[normalize_cell_text(c) for c in row] for row in table_data]
                    table_norm = repair_table_structure(table_norm)
                    
                    # Filter out empty or invalid tables
                    if len(table_norm) > 0 and any(any(cell for cell in row) for row in table_norm):
                        # Remove completely empty rows
                        table_norm = [row for row in table_norm if any(str(cell).strip() for cell in row)]
                        
                        if len(table_norm) > 0:
                            blocks.append({
                                "type": "table",
                                "y": float(y_mid),
                                "content": table_norm,
                                "strategy": strategy
                            })
                except Exception:
                    continue
            
            # 3) Extract images
            images = extract_images_from_page(page, pno)
            for img in images:
                blocks.append({
                    "type": "image",
                    "y": float(img['y']),
                    "content": f"[Image: {img['width']}x{img['height']}]"
                })
            
            # 4) Check if page is scanned (no text) and perform OCR
            if len(blocks) < 5 and OCR_AVAILABLE:
                ocr_text = perform_ocr_on_page(pdf_path, pno)
                if ocr_text and len(ocr_text.strip()) > 50:
                    blocks.append({
                        "type": "text",
                        "y": 100.0,
                        "content": f"[OCR Extracted Text]\n{ocr_text}"
                    })

            blocks_sorted = sorted(blocks, key=lambda b: b["y"])
            pages_out.append({
                "page_no": pno,
                "page_height": float(page_height),
                "page_width": float(page_width),
                "blocks": blocks_sorted
            })

    return pages_out


# ---------------------------
# Build Excel: one sheet per page; interleave text & tables; numeric conversion; freeze panes; y->rows mapping
# ---------------------------
def build_excel(pages, logo_path=None, merge_text_columns=10, target_rows_per_page=60, form_fields=None):
    """
    pages: output from extract_pdf_content
    merge_text_columns: number of columns to merge for text flow (A..J if 10)
    target_rows_per_page: controls vertical mapping resolution (higher -> finer alignment)
    """
    wb = Workbook()
    default = wb.active
    first_sheet = True

    for page in pages:
        if first_sheet:
            ws = default
            ws.title = f"Page_{page['page_no']}"
            first_sheet = False
        else:
            ws = wb.create_sheet(title=f"Page_{page['page_no']}")

        current_row = 1

        page_height = page.get("page_height", 792.0)
        rows_per_point = max(0.01, target_rows_per_page / page_height)

        # insert logo if present
        if logo_path and os.path.exists(logo_path):
            try:
                img = XLImage(logo_path)
                max_w = 300
                if getattr(img, "width", 0) > max_w:
                    ratio = max_w / img.width
                    img.width = int(img.width * ratio)
                    img.height = int(img.height * ratio)
                ws.add_image(img, "A1")
                current_row = 12
            except Exception:
                current_row = 1

        text_end_col = merge_text_columns
        
        # Add form fields at the top if present
        if form_fields and page['page_no'] == 1:
            ws.cell(row=current_row, column=1, value="Form Fields Detected:").font = Font(bold=True, size=12)
            current_row += 1
            for field in form_fields[:10]:  # Limit to first 10
                ws.cell(row=current_row, column=1, value=field['name'])
                ws.cell(row=current_row, column=2, value=field['value'])
                current_row += 1
            current_row += 1

        for block in page["blocks"]:
            # map block.y to a target row (1-indexed)
            target_row = 1 + int(block["y"] * rows_per_point)
            # ensure target_row at least current_row (monotonic)
            if target_row > current_row:
                # leave vertical gap
                current_row = target_row

            if block["type"] == "text":
                # Handle text with spatial positioning
                if isinstance(block["content"], list):  # Multi-column text
                    columns_data = block.get("columns", [])
                    if columns_data:
                        # Place text in appropriate columns
                        for col_idx, x_pos, text in columns_data:
                            excel_col = min(col_idx + 1, text_end_col)
                            cell = ws.cell(row=current_row, column=excel_col)
                            cell.value = text
                            cell.alignment = Alignment(
                                horizontal=block.get("alignment", "left"),
                                wrapText=True,
                                vertical="top"
                            )
                    current_row += 1
                else:
                    # Single text line with alignment
                    text = str(block["content"]).strip()
                    if not text:
                        continue
                    
                    # Determine starting column based on x position
                    x_start = block.get("x_start", 0)
                    page_width = page.get("page_width", 612.0)
                    
                    # Map x position to Excel column (1-10 range)
                    if x_start > page_width * 0.6:
                        start_col = max(1, text_end_col - 3)  # Right side
                    elif x_start > page_width * 0.3:
                        start_col = text_end_col // 2  # Center
                    else:
                        start_col = 1  # Left side
                    
                    # Merge cells for text flow
                    end_col = min(start_col + 5, text_end_col)
                    try:
                        ws.merge_cells(start_row=current_row, start_column=start_col, 
                                     end_row=current_row, end_column=end_col)
                    except Exception:
                        pass
                    
                    cell = ws.cell(row=current_row, column=start_col)
                    cell.value = text
                    
                    # Apply alignment from PDF
                    alignment_type = block.get("alignment", "left")
                    cell.alignment = Alignment(
                        horizontal=alignment_type,
                        wrapText=True,
                        vertical="top"
                    )
                    
                    if len(text) > 150:
                        try:
                            ws.row_dimensions[current_row].height = 30
                        except:
                            pass
                    current_row += 1

            elif block["type"] == "image":
                # Handle image placeholders
                text = block["content"]
                start_cell = ws.cell(row=current_row, column=1)
                start_cell.value = text
                start_cell.font = Font(italic=True, color="666666")
                start_cell.alignment = Alignment(wrapText=True, vertical="top")
                current_row += 1
            
            elif block["type"] == "table":
                table = block["content"]
                if not table:
                    continue
                nrows = len(table)
                ncols = max((len(row) for row in table), default=1)

                # Enhanced header detection (supports multi-row headers)
                header_rows = 0
                if nrows >= 2:
                    # Check first 2 rows for header patterns
                    for check_rows in range(min(3, nrows)):
                        row = table[check_rows]
                        non_empty = sum(1 for c in row if str(c).strip() != "")
                        if non_empty == 0:
                            break
                        
                        likely_header = 0
                        for cell in row:
                            parsed_num, is_num = try_parse_number(cell)
                            parsed_date, is_date = try_parse_date(cell)
                            if not is_num and not is_date and str(cell).strip():
                                likely_header += 1
                        
                        if non_empty > 0 and (likely_header / non_empty) >= 0.6:
                            header_rows = check_rows + 1
                        else:
                            break

                # Write rows with smart type conversion and formatting
                col_values = [[] for _ in range(ncols)]
                col_types = [set() for _ in range(ncols)]  # Track data types per column
                
                for r_idx, row_values in enumerate(table, start=0):
                    for c_idx in range(ncols):
                        val = ""
                        if c_idx < len(row_values):
                            val = row_values[c_idx]
                        vnorm = str(val).strip()
                        
                        # Smart type detection
                        parsed_date, is_date = try_parse_date(vnorm)
                        if is_date:
                            write_val = parsed_date
                            col_types[c_idx].add('date')
                        else:
                            parsed_num, is_num = try_parse_number(vnorm)
                            if is_num:
                                write_val = parsed_num
                                col_types[c_idx].add('number')
                            else:
                                write_val = vnorm
                                col_types[c_idx].add('text')
                        
                        cell = ws.cell(row=current_row + r_idx, column=1 + c_idx, value=write_val)
                        cell.alignment = Alignment(wrapText=True, vertical="top")
                        
                        # Apply formatting based on type
                        if is_date:
                            cell.number_format = 'MM/DD/YYYY'
                        elif is_num and isinstance(write_val, float) and write_val < 1:
                            # Percentage or decimal
                            if abs(write_val) < 0.01:
                                cell.number_format = '0.00%'
                            else:
                                cell.number_format = '#,##0.00'
                        elif is_num and isinstance(write_val, (int, float)):
                            # Regular number
                            if isinstance(write_val, int) or write_val == int(write_val):
                                cell.number_format = '#,##0'
                            else:
                                cell.number_format = '#,##0.00'
                        
                        col_values[c_idx].append(write_val)

                # Merges disabled to prevent scroll issues
                # merges = block.get("merges", [])
                
                # Style table with header highlighting
                _apply_table_style(ws, current_row, 1, nrows, ncols, header_rows=header_rows)
                
                # Add header background and enable auto-filter
                if header_rows > 0:
                    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                    for hr in range(header_rows):
                        for c in range(1, ncols + 1):
                            cell = ws.cell(row=current_row + hr, column=c)
                            cell.fill = header_fill
                            cell.font = Font(bold=True, size=11)
                    
                    # Add filter to header row
                    try:
                        end_col = get_column_letter(ncols)
                        ws.auto_filter.ref = f"A{current_row}:{end_col}{current_row + nrows - 1}"
                    except:
                        pass

                # REMOVED: Freeze panes can cause scroll issues - let users set manually if needed

                # Optimal column width with type-aware sizing
                for c_idx in range(ncols):
                    col_letter = get_column_letter(c_idx + 1)
                    # Adjust width based on column data type
                    if 'number' in col_types[c_idx] and len(col_types[c_idx]) == 1:
                        optimal_width = min(15, calculate_optimal_width(col_values[c_idx]))
                    elif 'date' in col_types[c_idx]:
                        optimal_width = 12
                    else:
                        optimal_width = calculate_optimal_width(col_values[c_idx])
                    ws.column_dimensions[col_letter].width = optimal_width
                
                current_row += nrows + 1

        # final auto-fit for left text area
        try:
            a_w = ws.column_dimensions["A"].width or 40
            ws.column_dimensions["A"].width = min(120, max(20, int(a_w)))
        except Exception:
            pass
        
        # Enable filters on first table if exists
        try:
            if ws.max_row > 1 and ws.max_column > 1:
                ws.auto_filter.ref = f"A1:{get_column_letter(min(ws.max_column, 20))}{min(ws.max_row, 1000)}"
        except:
            pass
        
        # Set print settings for better output
        try:
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToHeight = 0
            ws.page_setup.fitToWidth = 1
            ws.print_options.horizontalCentered = True
            ws.sheet_properties.pageSetUpPr.fitToPage = True
        except:
            pass

    return wb


# ---------------------------
# Flask routes
# ---------------------------

@app.route("/")
def index():
    # ensure templates/pdftoexcel.html exists
    return render_template("pdftoexcel.html")


@app.route("/upload", methods=["POST"])
def upload():
    if "pdfFile" not in request.files:
        flash("No file part")
        return redirect("/")

    file = request.files["pdfFile"]
    if file.filename == "":
        flash("No selected file")
        return redirect("/")

    if file and allowed(file.filename):
        pdf_name = secure_filename(file.filename)
        pdf_path = os.path.join(app.config["UPLOAD_FOLDER"], pdf_name)
        file.save(pdf_path)

        # Extract form fields if present
        form_fields = extract_form_fields(pdf_path)
        
        pages = extract_pdf_content(pdf_path)
        logo = extract_logo(pdf_path)
        wb = build_excel(pages, logo_path=logo, form_fields=form_fields)

        # stream workbook to user without writing a permanent file
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        # cleanup small files
        try:
            if logo and os.path.exists(logo):
                os.remove(logo)
            if os.path.exists(pdf_path):
                os.remove(pdf_path)
        except Exception:
            pass

        return send_file(
            stream,
            as_attachment=True,
            download_name="Converted.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    flash("Invalid file format")
    return redirect("/")


if __name__ == "__main__":
    # debug True for dev, change for prod
    app.run(debug=True)
